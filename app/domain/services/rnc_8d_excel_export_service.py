from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2]
    / "content"
    / "templates"
    / "quality"
    / "rnc_8d_template.xlsx"
)

SUPPLIER_BY_BRANCH = {
    "01": "12243 - Delpi Componentes Ltda EPP",
    "02": "12243 - Delpi Componentes Ltda EPP",
}

ANNEX_SHEET_CANDIDATES = ("Anexos(Evidencias)", "Anexos", "Attachment")
ANNEX_IMAGE_MAX_WIDTH_PX = 480
IMAGE_MIME_PREFIX = "image/"


def _excel_date(value: str | date | datetime | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _set(ws, cell: str, value: Any) -> None:
    if value is None:
        return
    text = str(value).strip() if not isinstance(value, (date, datetime, int, float)) else value
    if text == "":
        return
    ws[cell] = text


def is_image_evidence(evidence: dict[str, Any]) -> bool:
    mime = str(evidence.get("mime_type") or "").lower()
    if mime.startswith(IMAGE_MIME_PREFIX):
        return True
    evidence_type = str(evidence.get("type") or "").lower()
    return evidence_type == "image"


def _resolve_annex_sheet(wb) -> Any | None:
    for name in ANNEX_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _embed_annex_images(ws: Any, image_annexes: list[dict[str, Any]]) -> None:
    row = 3
    for item in image_annexes:
        content = item.get("content")
        if not content:
            continue
        try:
            img = XlImage(io.BytesIO(content))
        except Exception:
            continue
        if img.width > ANNEX_IMAGE_MAX_WIDTH_PX:
            ratio = ANNEX_IMAGE_MAX_WIDTH_PX / img.width
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, f"A{row}")
        label = str(item.get("file_name") or "").strip()
        description = str(item.get("description") or "").strip()
        if label:
            ws[f"D{row}"] = label
        if description:
            ws[f"D{row + 1}"] = description
        row += max(12, int(img.height / 15) + 2)


def _material_label(plan: dict[str, Any]) -> str:
    code = (plan.get("product_code") or "").strip()
    desc = (plan.get("product_description") or "").strip()
    if code and desc:
        return f"{code} - {desc}"
    return code or desc


def collect_image_annexes_for_export(
    *,
    plan_id: str,
    evidences: list[dict[str, Any]],
    storage: Any,
) -> list[dict[str, Any]]:
    annexes: list[dict[str, Any]] = []
    for evidence in evidences:
        if not is_image_evidence(evidence):
            continue
        stored_name = str(evidence.get("stored_name") or "").strip()
        if not stored_name:
            continue
        try:
            file_path = storage.resolve_file(plan_id=plan_id, stored_name=stored_name)
            annexes.append(
                {
                    "file_name": evidence.get("file_name") or stored_name,
                    "description": evidence.get("description"),
                    "content": file_path.read_bytes(),
                }
            )
        except Exception:
            continue
    return annexes


def build_rnc_8d_workbook(
    detail: dict[str, Any],
    *,
    image_annexes: list[dict[str, Any]] | None = None,
) -> bytes:
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"Template 8D não encontrado: {TEMPLATE_PATH}")

    plan = detail.get("plan") or {}
    payload = plan.get("template_payload") or {}
    if not isinstance(payload, dict):
        payload = {}

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["R8D"]

    nc = payload.get("nc_description") or {}
    classification = payload.get("classification") or {}
    effectiveness = payload.get("effectiveness") or {}
    preventive = payload.get("preventive") or {}
    containment_rows = payload.get("containment") or []
    documentation = payload.get("documentation_updates") or []
    team = detail.get("team_members") or []
    five_whys = detail.get("five_whys") or {}
    actions = detail.get("actions") or []

    branch = plan.get("branch_code") or "01"
    _set(ws, "I4", plan.get("client_nc_registry"))
    _set(ws, "E5", payload.get("supplier_name") or SUPPLIER_BY_BRANCH.get(branch, SUPPLIER_BY_BRANCH["01"]))
    _set(ws, "E6", _material_label(plan))
    _set(ws, "E7", payload.get("material_specification"))
    _set(ws, "E8", payload.get("purchase_order"))
    _set(ws, "E9", payload.get("invoice_number"))
    invoice_date = _excel_date(payload.get("invoice_date"))
    if invoice_date:
        ws["E10"] = invoice_date
    _set(ws, "E11", payload.get("defective_quantity"))
    _set(ws, "E12", payload.get("return_invoice_number"))
    _set(ws, "J5", plan.get("customer_contact"))
    _set(ws, "J6", payload.get("contact_phone"))
    _set(ws, "J7", payload.get("contact_fax"))
    _set(ws, "J8", payload.get("client_batch"))
    _set(ws, "J9", plan.get("batch_number"))
    _set(ws, "J10", payload.get("batch_quantity"))
    _set(ws, "J11", payload.get("disposition"))
    _set(ws, "J12", payload.get("rejected_quantity"))

    if classification.get("end_customer"):
        _set(ws, "K3", plan.get("customer_name"))
    report_date = _excel_date(payload.get("report_date") or plan.get("reported_at"))
    if report_date:
        ws["K1"] = report_date

    _set(ws, "A15", nc.get("characteristic") or payload.get("nc_characteristic"))
    _set(ws, "E15", nc.get("specified"))
    verified = nc.get("verified") or plan.get("reported_problem")
    _set(ws, "I15", verified)
    _set(ws, "C18", nc.get("observations") or payload.get("observations"))

    return_by = _excel_date(payload.get("return_by"))
    if return_by:
        ws["D21"] = return_by
    _set(ws, "G21", payload.get("attention_to"))
    _set(ws, "J21", payload.get("attention_email"))

    leader = next((m for m in team if m.get("is_leader")), team[0] if team else None)
    members = [m for m in team if not m.get("is_leader")]
    if leader:
        _set(ws, "D23", leader.get("member_name"))
        _set(ws, "H23", leader.get("department"))
    member_rows = [25, 26, 27, 28]
    for index, member in enumerate(members[:4]):
        row = member_rows[index]
        _set(ws, f"D{row}", member.get("member_name"))
        _set(ws, f"H{row}", member.get("department"))

    area_row_map = {
        "end_customer": 35,
        "client": 35,
        "client_plant": 37,
        "supplier": 39,
    }
    for item in containment_rows:
        area_key = (item.get("area") or "").lower()
        row = area_row_map.get(area_key)
        if not row:
            continue
        _set(ws, f"E{row}", item.get("quantity"))
        _set(ws, f"G{row}", item.get("action_plan"))
        _set(ws, f"J{row}", item.get("responsible"))
        containment_date = _excel_date(item.get("date"))
        if containment_date:
            ws[f"M{row}"] = containment_date

    occurrence_cols = ["E", "G", "I", "K", "M"]
    for index, col in enumerate(occurrence_cols):
        key = f"why_{index + 1}"
        _set(ws, f"{col}44", five_whys.get(key))

    detection_cols = ["E", "G", "I", "K", "M"]
    for index, col in enumerate(detection_cols):
        key = f"detection_why_{index + 1}"
        _set(ws, f"{col}49", five_whys.get(key))

    corrective_actions = [
        a for a in actions if a.get("action_type") == "corrective" or a.get("cause_track")
    ]
    action_rows = [56, 57, 58, 59, 60]
    for index, action in enumerate(corrective_actions[:5]):
        row = action_rows[index]
        track = action.get("cause_track")
        if track == "occurrence":
            _set(ws, f"D{row}", "Ocorrência")
        elif track == "detection":
            _set(ws, f"D{row}", "Detecção")
        _set(ws, f"F{row}", action.get("description"))
        _set(ws, f"J{row}", action.get("responsible_name"))
        due = _excel_date(action.get("due_date"))
        if due:
            ws[f"M{row}"] = due

    resolved = effectiveness.get("resolved_how") or plan.get("effectiveness_notes")
    _set(ws, "D64", resolved)
    _set(ws, "D71", effectiveness.get("ok_material_date"))
    _set(ws, "F71", effectiveness.get("new_parts_identification"))
    _set(ws, "J71", effectiveness.get("verification_responsible"))
    verification_date = _excel_date(effectiveness.get("verification_date"))
    if verification_date:
        ws["L71"] = verification_date

    _set(ws, "D73", preventive.get("how_avoid_future"))
    _set(ws, "D77", preventive.get("other_processes_products"))
    _set(ws, "D84", preventive.get("evaluation_responsible"))
    evaluation_date = _excel_date(preventive.get("evaluation_completion_date"))
    if evaluation_date:
        ws["I84"] = evaluation_date

    doc_rows = [86, 87, 88, 89]
    for index, doc in enumerate(documentation[:4]):
        row = doc_rows[index]
        _set(ws, f"F{row}", doc.get("document"))
        _set(ws, f"I{row}", doc.get("responsible"))
        doc_date = _excel_date(doc.get("date"))
        if doc_date:
            ws[f"K{row}"] = doc_date

    closure = payload.get("client_closure_note")
    _set(ws, "D108", closure or "Conforme status da nota QM.")

    annex_ws = _resolve_annex_sheet(wb)
    if annex_ws is not None and image_annexes:
        _embed_annex_images(annex_ws, image_annexes)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
